from __future__ import annotations

from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
import sqlite3
from typing import Any, Callable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .exceptions import ImportExecutionError
from .link_parser import parse_mega_links
from .models import ImportRowError, ImportStats, WorkRow
from .normalizers import (
    is_iso_sale_date,
    normalize_optional_text,
    normalize_sale_date,
    normalize_work_code,
    parse_circles,
    parse_voice_actors,
)
from .repositories import (
    add_import_error,
    create_import,
    finish_import,
    refresh_circle_mappings,
    refresh_voice_actor_mappings,
    replace_active_links_if_changed,
    upsert_work,
)
from .sizes import parse_size_bytes

REQUIRED_HEADERS = (
    "RJcode",
    "标签",
    "MEGA链接",
    "销售日期",
    "声优",
    "标题",
    "备注",
    "类型",
    "社团",
    "档案大小",
    "MP3大小",
)

ImportProgressCallback = Callable[[int, int | None], None]


@dataclass(frozen=True)
class ImportResult:
    import_id: int
    source_path: Path
    database_path: Path | None
    stats: ImportStats
    status: str
    notes: str | None = None
    error_log_path: Path | None = None


def import_excel_workbook(
    conn: sqlite3.Connection,
    xlsx_path: Path,
    *,
    database_path: Path | None = None,
    progress_callback: ImportProgressCallback | None = None,
) -> ImportResult:
    xlsx_path = Path(xlsx_path).expanduser()
    if not xlsx_path.exists():
        raise ImportExecutionError(f"Import failed: {xlsx_path} does not exist.")
    if not xlsx_path.is_file():
        raise ImportExecutionError(f"Import failed: {xlsx_path} is not a file.")

    try:
        workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises several concrete exception types.
        raise ImportExecutionError(f"Import failed: could not open {xlsx_path}: {exc}") from exc

    try:
        worksheet = workbook.worksheets[0]
        headers = _header_map(worksheet)
        missing_headers = [header for header in REQUIRED_HEADERS if header not in headers]
        if missing_headers:
            joined = ", ".join(missing_headers)
            raise ImportExecutionError(f"Import failed: missing required columns: {joined}")

        stats = ImportStats()
        now = _utc_now()
        import_id = create_import(conn, xlsx_path, now)
        conn.commit()

        try:
            conn.execute("BEGIN")
            _import_rows(
                conn,
                import_id,
                worksheet,
                headers,
                stats,
                now,
                progress_callback=progress_callback,
            )
            conn.commit()
        except Exception as exc:
            conn.rollback()
            notes = f"failed: {exc}"
            finish_import(conn, import_id, "failed", stats, notes)
            conn.commit()
            if isinstance(exc, ImportExecutionError):
                raise
            raise ImportExecutionError(f"Import failed: {exc}") from exc

        notes = "completed"
        finish_import(conn, import_id, "completed", stats, notes)
        conn.commit()
        return ImportResult(
            import_id=import_id,
            source_path=xlsx_path,
            database_path=database_path,
            stats=stats,
            status="completed",
            notes=notes,
        )
    finally:
        workbook.close()


def _import_rows(
    conn: sqlite3.Connection,
    import_id: int,
    worksheet: Worksheet,
    headers: dict[str, int],
    stats: ImportStats,
    now: str,
    *,
    progress_callback: ImportProgressCallback | None = None,
) -> None:
    total_rows = _worksheet_data_row_count(worksheet)
    if progress_callback is not None:
        progress_callback(0, total_rows)

    scanned_rows = 0
    for row_number, values in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        scanned_rows += 1
        if progress_callback is not None:
            progress_callback(scanned_rows, total_rows)

        if _is_empty_row(values):
            continue

        stats.total_rows += 1
        row_data = _row_data(headers, values)
        work_code = _try_work_code(row_data.get("RJcode"))
        if work_code is None:
            _record_error(
                conn,
                import_id,
                stats,
                ImportRowError(
                    error_type="missing_work_code",
                    message="RJcode is required.",
                    row_number=row_number,
                    raw_value=_raw_value(row_data.get("RJcode")),
                ),
                now,
            )
            continue

        work_row, row_errors = _build_work_row(row_data, row_number, work_code)
        for error in row_errors:
            _record_error(conn, import_id, stats, error, now)

        work_id, inserted, metadata_changed = upsert_work(conn, work_row, now)
        refresh_voice_actor_mappings(conn, work_id, work_row.voice_actor_names)
        refresh_circle_mappings(conn, work_id, work_row.circle_names)

        link_result = parse_mega_links(
            row_data.get("MEGA链接"),
            row_number=row_number,
            work_code=work_code,
        )
        for error in link_result.errors:
            _record_error(conn, import_id, stats, error, now)

        has_invalid_root_json = any(
            error.error_type == "invalid_mega_json" for error in link_result.errors
        )
        link_set_changed = False
        if not has_invalid_root_json:
            link_set_changed = replace_active_links_if_changed(
                conn,
                work_id,
                link_result.links,
                now,
            )

        if inserted:
            stats.inserted_works += 1
        elif metadata_changed:
            stats.updated_works += 1
        elif not link_set_changed:
            stats.skipped_works += 1

        if link_set_changed:
            stats.link_sets_changed += 1


def _build_work_row(
    row_data: dict[str, Any],
    row_number: int,
    work_code: str,
) -> tuple[WorkRow, tuple[ImportRowError, ...]]:
    errors: list[ImportRowError] = []
    sale_date = normalize_sale_date(row_data.get("销售日期"))
    if not is_iso_sale_date(sale_date):
        errors.append(
            ImportRowError(
                error_type="invalid_sale_date_format",
                message="Sale date is not in YYYY-MM-DD format.",
                row_number=row_number,
                work_code=work_code,
                raw_value=_raw_value(row_data.get("销售日期")),
            )
        )

    archive_size_raw = normalize_optional_text(row_data.get("档案大小"))
    archive_size_bytes = parse_size_bytes(row_data.get("档案大小"))
    if archive_size_raw is not None and archive_size_bytes is None:
        errors.append(
            ImportRowError(
                error_type="invalid_size",
                message="Archive size could not be parsed.",
                row_number=row_number,
                work_code=work_code,
                raw_value=archive_size_raw,
            )
        )

    mp3_size_raw = normalize_optional_text(row_data.get("MP3大小"))
    mp3_size_bytes = parse_size_bytes(row_data.get("MP3大小"))
    if mp3_size_raw is not None and mp3_size_bytes is None:
        errors.append(
            ImportRowError(
                error_type="invalid_size",
                message="MP3 size could not be parsed.",
                row_number=row_number,
                work_code=work_code,
                raw_value=mp3_size_raw,
            )
        )

    voice_actor_raw = normalize_optional_text(row_data.get("声优"))
    circle_raw = normalize_optional_text(row_data.get("社团"))
    work_row = WorkRow(
        work_code=work_code,
        title=normalize_optional_text(row_data.get("标题")),
        tags_raw=normalize_optional_text(row_data.get("标签")),
        sale_date=sale_date,
        voice_actor_raw=voice_actor_raw,
        voice_actor_names=parse_voice_actors(voice_actor_raw),
        note=normalize_optional_text(row_data.get("备注")),
        work_type=normalize_optional_text(row_data.get("类型")),
        circle_raw=circle_raw,
        circle_names=parse_circles(circle_raw),
        archive_size_raw=archive_size_raw,
        archive_size_bytes=archive_size_bytes,
        mp3_size_raw=mp3_size_raw,
        mp3_size_bytes=mp3_size_bytes,
        source_row_number=row_number,
    )
    return work_row, tuple(errors)


def _header_map(worksheet: Worksheet) -> dict[str, int]:
    first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if first_row is None:
        return {}

    headers: dict[str, int] = {}
    for index, value in enumerate(first_row):
        header = normalize_optional_text(value)
        if header is not None and header not in headers:
            headers[header] = index
    return headers


def _row_data(headers: dict[str, int], values: tuple[Any, ...]) -> dict[str, Any]:
    return {
        header: values[index] if index < len(values) else None
        for header, index in headers.items()
    }


def _try_work_code(value: Any) -> str | None:
    try:
        return normalize_work_code(value)
    except ValueError:
        return None


def _record_error(
    conn: sqlite3.Connection,
    import_id: int,
    stats: ImportStats,
    error: ImportRowError,
    now: str,
) -> None:
    add_import_error(conn, import_id, error, now)
    stats.error_count += 1


def _raw_value(value: Any) -> str | None:
    if value is None:
        return None
    return str(value)


def _is_empty_row(values: tuple[Any, ...]) -> bool:
    return all(normalize_optional_text(value) is None for value in values)


def _worksheet_data_row_count(worksheet: Worksheet) -> int | None:
    max_row = worksheet.max_row
    if max_row is None:
        return None
    return max(max_row - 1, 0)


def _utc_now() -> str:
    return datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")
